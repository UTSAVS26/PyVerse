import os
import csv
import re
import logging
import openpyxl
from pathlib import Path
from typing import Union
import argparse
from openpyxl.utils.exceptions import InvalidFileException

def excel_to_csv(input_file: Union[str, Path], output_folder: Union[str, Path] = "output") -> list[str]:
    """
    Convert an Excel file to CSV files, one per sheet.

    Args:
        input_file (str | Path): Path to the input Excel file (.xlsx).
        output_folder (str | Path, optional): Directory to save output CSV files. Defaults to "output".

    Returns:
        list[str]: List of created CSV file paths
    """
    os.makedirs(output_folder, exist_ok=True)
    generated_csvs = []

    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
    except (FileNotFoundError, InvalidFileException) as exc:
        logging.error("unable to open workbook %s: %s", input_file, exc)
        return []

    for sheet_name in wb.sheetnames:
        try:
            sheet = wb[sheet_name]
            safe_title = re.sub(r'[^A-Za-z0-9._-]', "_", sheet.title)
            csv_name = f"{Path(input_file).stem}_{safe_title}.csv"
            csv_path = os.path.join(output_folder, csv_name)

            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(row)

            generated_csvs.append(csv_path)

        except Exception as e:
            logging.warning("error processing sheet '%s' in '%s': %s", sheet_name, input_file, e)

    wb.close()
    return generated_csvs


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert an Excel file to CSV.")
    parser.add_argument("input_file", help="Path to the .xlsx file")
    parser.add_argument("-o", "--output", default="output", help="Destination directory for CSV files")
    args = parser.parse_args()

    results = excel_to_csv(args.input_file, args.output)
    if results:
        print("csv files generated:")
        for path in results:
            print("-", path)
    else:
        print("no files were generated.")

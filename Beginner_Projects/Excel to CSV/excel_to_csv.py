import csv
import re
import logging
import argparse
from pathlib import Path
from typing import Union
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

def excel_to_csv(input_file: Union[str, Path], output_folder: Union[str, Path] = "output") -> list[str]:
    """
    Convert an Excel file to CSV files, one per sheet.

    Args:
        input_file (str | Path): Path to the input Excel file (.xlsx).
        output_folder (str | Path, optional): Directory to save output CSV files. Defaults to "output".

    Returns:
        list[str]: List of created CSV file paths
    """
    input_path = Path(input_file)
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)
    generated_csvs = []

    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except (FileNotFoundError, PermissionError, InvalidFileException) as exc:
        logging.error("unable to open workbook %s: %s", input_path, exc)
        return []

    for idx, sheet_name in enumerate(wb.sheetnames, start=1):
        try:
            sheet = wb[sheet_name]
            safe_title = re.sub(r'[^A-Za-z0-9._-]', "_", sheet.title)
            if not safe_title:
                safe_title = f"Sheet{idx}"

            csv_name = f"{input_path.stem}_{safe_title}.csv"
            csv_path = output_path / csv_name
            counter = 1
            while csv_path.exists():
                csv_path = output_path / f"{input_path.stem}_{safe_title}_{counter}.csv"
                counter += 1

            with csv_path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow("" if cell is None else cell for cell in row)

            generated_csvs.append(str(csv_path))

        except (PermissionError, OSError, csv.Error) as exc:
            logging.warning("error processing sheet '%s' in '%s': %s", sheet_name, input_path, exc)
        except Exception:
            logging.exception("unexpected error processing sheet '%s'", sheet_name)

    wb.close()
    return generated_csvs

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    parser = argparse.ArgumentParser(description="Convert an Excel file to CSV.")
    parser.add_argument("input_file", help="Path to the .xlsx file")
    parser.add_argument("-o", "--output", default="output", help="Destination directory for CSV files")
    args = parser.parse_args()

    results = excel_to_csv(args.input_file, args.output)
    if results:
        print("csv files generated:")
        for path in results:
            print("-", path)
    else:
        print("no files were generated.")
